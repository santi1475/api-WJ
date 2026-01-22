import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import Cliente

def generar_excel_masivo(lista_rucs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Clientes"

    # fuentes
    font_header = Font(name='Calibri', size=11, bold=True, color='000000') # Negrita
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # bordes
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # colores
    fill_blue_light = PatternFill(start_color='C5D9F1', fill_type='solid') 
    fill_yellow = PatternFill(start_color='FFFF00', fill_type='solid')    
    fill_orange_light = PatternFill(start_color='FDE9D9', fill_type='solid') 
    fill_teal = PatternFill(start_color='92CDDC', fill_type='solid')       
    fill_pink = PatternFill(start_color='F2DCDB', fill_type='solid')       
    fill_purple = PatternFill(start_color='E4DFEC', fill_type='solid')     

    # colores condicionales 
    cat_fill_a = PatternFill(start_color='C6EFCE', fill_type='solid') 
    cat_fill_b = PatternFill(start_color='FFEB9C', fill_type='solid')
    cat_fill_c = PatternFill(start_color='FFC7CE', fill_type='solid')


    # Lista
    headers = [
        'RUC', 'Razón Social', 'Propietario', 'DNI Propietario', 'Fecha Ingreso',
        'Estado', 'Código Control', 'Responsable', 
        'Régimen Tributario', 'Tipo Empresa', 'Categoría',
        'Régimen Laboral Tipo', 'Régimen Laboral Fecha',
        'Ingresos Mensuales', 'Ingresos Anuales', 'Libros Societarios', 'Selectivo Consumo',
        'PE', 
        'SOL Usuario', 'SOL Clave',
        'Detracción Cuenta', 'Detracción Usuario', 'Detracción Clave',
        'INEI Usuario', 'INEI Clave',
        'AFP Net Usuario', 'AFP Net Clave',
        'Viva Essalud Usuario', 'Viva Essalud Clave',
        'SIS Usuario', 'SIS Clave',
        'Clave OSCE', 'Clave SENCICO'
    ]


    header_color_map = {
        'RUC': fill_yellow,
        'SOL Usuario': fill_yellow,
        'SOL Clave': fill_yellow,
        'Responsable': fill_yellow,

        'Razón Social': fill_blue_light,
        'Propietario': fill_blue_light,
        'DNI Propietario': fill_blue_light,
        'Fecha Ingreso': fill_blue_light,
        'Código Control': fill_blue_light,
        'Detracción Cuenta': fill_blue_light,
        'Detracción Usuario': fill_blue_light,
        'Detracción Clave': fill_blue_light,
        'INEI Usuario': fill_blue_light,
        'INEI Clave': fill_blue_light,
        'Clave OSCE': fill_blue_light,
        'Clave SENCICO': fill_blue_light,
        'Régimen Tributario': fill_blue_light,
        'Régimen Laboral Tipo': fill_blue_light,
        'Régimen Laboral Fecha': fill_blue_light,
        'AFP Net Usuario': fill_blue_light,
        'AFP Net Clave': fill_blue_light,
        'Viva Essalud Usuario': fill_blue_light,
        'Viva Essalud Clave': fill_blue_light,
        'SIS Usuario': fill_blue_light,
        'SIS Clave': fill_blue_light,
        'PE': fill_blue_light,
        'Ingresos Mensuales': fill_orange_light,
        'Ingresos Anuales': fill_orange_light,

        'Libros Societarios': fill_teal,

        'Selectivo Consumo': fill_pink,

        'Categoría': fill_purple,
    }

    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
        
        column_name = cell.value
        if column_name in header_color_map:
            cell.fill = header_color_map[column_name]

    clientes = Cliente.objects.filter(pk__in=lista_rucs).select_related('credenciales', 'responsable')

    for cliente in clientes:
        try:
            creds = cliente.credenciales
        except:
            creds = None

        fila = [
            cliente.ruc,
            cliente.razon_social,
            cliente.propietario,
            cliente.dni_propietario,
            cliente.fecha_ingreso,
            'Activo' if cliente.estado else 'Inactivo',
            cliente.codigo_control,
            str(cliente.responsable) if cliente.responsable else 'Sin Asignar',
            cliente.get_regimen_tributario_display(),
            cliente.get_tipo_empresa_display(),
            cliente.get_categoria_display(),
            cliente.regimen_laboral_tipo,
            cliente.regimen_laboral_fecha,
            cliente.ingresos_mensuales,
            cliente.ingresos_anuales,
            cliente.libros_societarios,
            'Sí' if cliente.selectivo_consumo else 'No',
            creds.pe if creds else '',
            creds.sol_usuario if creds else '', creds.sol_clave if creds else '',
            creds.detraccion_cuenta if creds else '', creds.detraccion_usuario if creds else '', creds.detraccion_clave if creds else '',
            creds.inei_usuario if creds else '', creds.inei_clave if creds else '',
            creds.afp_net_usuario if creds else '', creds.afp_net_clave if creds else '',
            creds.viva_essalud_usuario if creds else '', creds.viva_essalud_clave if creds else '',
            creds.sis_usuario if creds else '', creds.sis_clave if creds else '',
            creds.clave_osce if creds else '',
            creds.clave_sencico if creds else ''
        ]
        
        ws.append(fila)
        
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = align_center #

            header_value = headers[cell.column - 1] 
            
            if header_value == 'Categoría' and cell.value:
                valor_texto = str(cell.value).upper() 
                if 'CATEGORÍA A' in valor_texto or cell.value == 'A':
                    cell.fill = cat_fill_a
                elif 'CATEGORÍA B' in valor_texto or cell.value == 'B':
                    cell.fill = cat_fill_b
                elif 'CATEGORÍA C' in valor_texto or cell.value == 'C':
                    cell.fill = cat_fill_c

    for column_cells in ws.columns:
        length = 0
        column_letter = column_cells[0].column_letter
        
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > length:
                        length = cell_length
            except:
                pass
        
        adjusted_width = (length + 2) * 1.1
        ws.column_dimensions[column_letter].width = adjusted_width

    return wb